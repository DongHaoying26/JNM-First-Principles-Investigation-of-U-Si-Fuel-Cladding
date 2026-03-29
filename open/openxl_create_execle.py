# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Create a new workbook
new_workbook = Workbook()

# Create Sheet1 and fill in data
sheet1 = new_workbook.active
sheet1.title = "Sheet1"

# Create Sheet2 and fill in mu, msi, mc data
sheet2 = new_workbook.create_sheet(title="Sheet2")
sheet2_data = [
    ['mu', dμ, dμ, dμ, dμ],
    ['msi', dμ, dμ, dμ, dμ],
    ['mc', dμ, dμ, dμ, dμ]
]

# Fill Sheet2 data and apply steel blue fill
steel_blue_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
for row in sheet2_data:
    sheet2.append(row)

for row in sheet2.iter_rows(min_row=1, max_row=3, min_col=1, max_col=5):
    for cell in row:
        cell.fill = steel_blue_fill

# Manually defined original data
sheet1_data_part1 = [
    ['Three-Phase-Equilibria'],
    ['Species', 'Fuel-cladding-product', 'Fuel-cladding-product', 'Fuel-cladding-product', 'Fuel-cladding-product'],
    ['U1vac', formatin energy],
    ['U2vac', formatin energy],
    ['Sivac', formatin energy],
    ['SionU1', formatin energy],
    ['SionU2', formatin energy],
    ['UonSi', formatin energy],
   
]

# Define fill color
dark_sea_green_fill = PatternFill(start_color="8FBC8F", end_color="8FBC8F", fill_type="solid")

# Fill in the first part of data
for row in sheet1_data_part1:
    sheet1.append(row)

# Copy the first part of data to columns C3 to E23
for row in range(3, 25):
    value = sheet1.cell(row=row, column=2).value
    sheet1.cell(row=row, column=3, value=value)
    sheet1.cell(row=row, column=4, value=value)
    sheet1.cell(row=row, column=5, value=value)

# Fill in the second part of data and reference values from Sheet2, apply steel blue fill
second_part_headers = ['mu', 'msi', 'mc']
for i, header in enumerate(second_part_headers):
    for j in range(5):
        sheet1.cell(row=i + 24, column=j + 1, value=header if j == 0 else f"=Sheet2!{chr(65+j)}{i+1}")
        sheet1.cell(row=i + 24, column=j + 1).fill = steel_blue_fill

# Initialize calculation results list
calculation_results = []

# Fill in calculation formulas and compute results
for i, row in enumerate(sheet1_data_part1[2:]):
    input_name = row[0]  # Get input data name
    input_value = row[1]

    # Initialize result variables
    results = []

    # Adjust calculation formulas as needed
    for j in range(4):  # Adjust to 4 to cover all Sheet2 data
        mu_value_col = chr(66 + j)
        msi_value_col = chr(66 + j)
        mc_value_col = chr(66 + j)
        mu_value = f"Sheet2!{mu_value_col}1"
        msi_value = f"Sheet2!{msi_value_col}2"
        mc_value = f"Sheet2!{mc_value_col}3"

        formula = None

        if input_name ==   'U1vac':
            formula = f"={input_value} - {mu_value}"
        elif input_name == 'U2vac':
            formula = f"={input_value} - {mu_value}"
        elif input_name == 'Sivac':
            formula = f"={input_value} - {msi_value}"
        elif input_name == 'SionU1':
            formula = f"={input_value} + {msi_value} - {mu_value}"
        elif input_name == 'SionU2':
            formula = f"={input_value} + {msi_value} - {mu_value}"
        elif input_name == 'UonSi':
            formula = f"={input_value} - {msi_value} + {mu_value}"



        results.append(formula)

    # Create calculation result row
    formula_row = [input_name] + results

    # Add to calculation results list
    calculation_results.append(formula_row)

# Write calculation results to the third part of the table in the same column, apply dark sea green fill
start_row = 27  # Calculation results start row
for i, row in enumerate(calculation_results):
    for j in range(5):
        cell = sheet1.cell(row=start_row + i, column=j + 1, value=row[j])
        cell.fill = dark_sea_green_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Center align all cell contents
for row in sheet1.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Create Sheet3 and fill in data
sheet3 = new_workbook.create_sheet(title="Sheet3")

# First part of data
sheet3_data_part1 = [
    ['Data Statistics'],
    ['E0', total energy],
    ['U1vac', total energy],
    ['U2vac', total energy],
    ['Sivac', total energy],
    ['SionU1', total energy],
    ['SionU2', total energy],
    ['UonSi', total energy],
 ]

# Fill in the first part of data
for row in sheet3_data_part1:
    sheet3.append(row)

# Second part of data
sheet3_data_part2 = [
    ['Mu', μ],
    ['Msi', μ],
    ['Mc', μ]
]

# Fill in the second part of data and apply cornflower blue fill
cornflower_blue_fill = PatternFill(start_color="6495ED", end_color="6495ED", fill_type="solid")
for row in sheet3_data_part2:
    sheet3.append(row)

for row in sheet3.iter_rows(min_row=24, max_row=26, min_col=1, max_col=2):
    for cell in row:
        cell.fill = cornflower_blue_fill

# Initialize calculation results list
calculation_results_sheet3 = []

# Fill in calculation formulas and compute results
for i, row in enumerate(sheet3_data_part1[2:]):
    input_name = row[0]  # Get input data name
    input_value = row[1]

    # Initialize result variables
    results = []

    # Adjust calculation formulas as needed
    for j in range(2):  # Adjust to 2 to cover all Sheet3 second part data
        mu_value_col = chr(66 + j)
        msi_value_col = chr(66 + j)
        mc_value_col = chr(66 + j)
        m0_value_col = chr(66 + j)
        mu_value = f"Sheet3!{mu_value_col}24"
        msi_value = f"Sheet3!{msi_value_col}25"
        mc_value = f"Sheet3!{mc_value_col}26"
        m0_value = f"Sheet3!{mc_value_col}2"

        formula = None

        if input_name == 'U1vac':
            formula = f"={input_value} - {mu_value} - {m0_value}"
        elif input_name == 'U2vac':
            formula = f"={input_value} - {mu_value} - {m0_value}"
        elif input_name == 'Sivac':
            formula = f"={input_value} - {msi_value} - {m0_value}"
        elif input_name == 'SionU1':
            formula = f"={input_value} + {msi_value} - {mu_value} - {m0_value}"
        elif input_name == 'SionU2':
            formula = f"={input_value} + {msi_value} - {mu_value} - {m0_value}"
        elif input_name == 'UonSi':
            formula = f"={input_value} - {msi_value} + {mu_value} - {m0_value}"


        results.append(formula)

    # Create calculation result row
    formula_row = [input_name] + results

    # Add to calculation results list
    calculation_results_sheet3.append(formula_row)

# Write calculation results to the third part of the table in the same column, apply dark sea green fill
start_row = 27  # Calculation results start row
for i, row in enumerate(calculation_results_sheet3):
    for j in range(2):
        cell = sheet3.cell(row=start_row + i, column=j + 1, value=row[j])
        cell.fill = dark_sea_green_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Center align all cell contents
for row in sheet3.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust column width to fit content
for col in sheet1.columns:
    max_length = 0
    column = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet1.column_dimensions[column].width = adjusted_width

for col in sheet3.columns:
    max_length = 0
    column = col[0].column_letter  # Get column letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet3.column_dimensions[column].width = adjusted_width

# Merge the first 5 columns of the first row in Sheet1
sheet1.merge_cells('A1:E1')
# Merge the first 2 columns of the first row in Sheet3
sheet3.merge_cells('A1:B1')

# Set format for merged cells
sheet1['A1'].font = Font(bold=True, size=sheet1['A1'].font.size + 5)
sheet3['A1'].font = Font(bold=True, size=sheet3['A1'].font.size + 5)

# Save the new workbook
new_file_path = 'Fuel_cladding.xlsx'
new_workbook.save(new_file_path)

print(f"New workbook created and saved as {new_file_path}")
